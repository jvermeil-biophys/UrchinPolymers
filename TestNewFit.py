# -*- coding: utf-8 -*-
"""
Created on Wed Nov 19 11:01:34 2025

@author: Utilisateur
"""

# translate_matlab_to_python.py

#%% --- 0. Imports

import os
import warnings
import numpy as np
from scipy import io
from scipy.optimize import minimize
import matplotlib.pyplot as plt
import pandas as pd

#%% --- 0. Settings (edit these to match your system) ---
root_path = r'Z:\Maribel\Paper Extracts\20250326_100CE\10x_100CE_2fps_DIC__5\D2-m5_10X_14-19'
filename = 'output_D2-m5_10X_14-19'   # .mat filename without extension

mat_path = os.path.join(root_path, filename + '.mat')

#%% --- 1. Load .mat and safely extract variables (with warnings) ---
if not os.path.isfile(mat_path):
    raise FileNotFoundError(f"MAT file not found: {mat_path}")

mat = io.loadmat(mat_path, squeeze_me=True, struct_as_record=False)

def safe_get(mat_dict, key, default=None):
    if key in mat_dict:
        return mat_dict[key]
    else:
        warnings.warn(f"'{key}' does not exist in the MAT file.")
        return default

# direct variables
t = safe_get(mat, 'tpulling', None)
k_var = safe_get(mat, 'k', None)
gamma1 = safe_get(mat, 'gamma1', None)
gamma2 = safe_get(mat, 'gamma2', None)

# nested dx.pulling_n handling (MATLAB struct -> numpy structured)
y_noisy = None
if 'dx' in mat:
    dx = mat['dx']
    # dx might be a numpy object array or a Matlab struct-like record
    try:
        y_noisy = getattr(dx, 'pulling_n')
    except Exception:
        # maybe dict-like
        if isinstance(dx, dict) and 'pulling_n' in dx:
            y_noisy = dx['pulling_n']
        else:
            warnings.warn("dx.pulling_n does not exist in the MAT file.")
            y_noisy = None
else:
    warnings.warn("'dx' does not exist in the MAT file.")
    y_noisy = None

# Convert t and y_noisy to numpy arrays if they are present
if t is None:
    raise ValueError("tpulling is required but not found in the .mat file.")
t = np.asarray(t, dtype=float).ravel()

if y_noisy is None:
    raise ValueError("dx.pulling_n is required but not found in the .mat file.")
y_noisy = np.asarray(y_noisy, dtype=float).ravel()

# --- 2. Model definition (constrained gamma2 >= 3*gamma1)
# parameterization: p = [log(k), log(gamma1), log(r)]
def model_constrained(p, t_arr):
    # p in log-space
    k = np.exp(p[0])
    gamma1_p = np.exp(p[1])
    r = np.exp(p[2])
    gamma2_p = 3.0 * gamma1_p + r**2
    # f(t) = (1/k)*(1 - exp(-k*t/gamma1)) + t/gamma2
    return (1.0 / k) * (1.0 - np.exp(-k * t_arr / gamma1_p)) + t_arr / gamma2_p

# --- initial values (cleaning & ensure positivity) ---
# fallback to small positive numbers if MATLAB variables missing
k_init = np.abs(k_var) if (k_var is not None and np.isfinite(k_var)) else 1e-6
k_init = max(k_init, 1e-6)
gamma1_init = np.abs(gamma1) if (gamma1 is not None and np.isfinite(gamma1)) else 1e-6
gamma1_init = max(gamma1_init, 1e-6)
# gamma2_init must be at least 3*gamma1_init + tiny
gamma2_init = np.abs(gamma2) if (gamma2 is not None and np.isfinite(gamma2)) else (3*gamma1_init + 1e-6)
gamma2_init = max(gamma2_init, 3*gamma1_init + 1e-6)
# derive r_init
r_init = np.sqrt(max(gamma2_init - 3*gamma1_init, 1e-6))

params0 = np.log([k_init, gamma1_init, r_init])  # initial guess in log-space

# --- 3. Nonlinear fit via Nelder-Mead (mimic fminsearch) ---
def cost_fun(p):
    pred = model_constrained(p, t)
    res = y_noisy - pred
    return np.sum(res**2)

res_opt = minimize(cost_fun, params0, method='Nelder-Mead', options={'maxiter': 10000, 'disp': False})
params_fit_log = res_opt.x
params_fit = np.exp(params_fit_log)
k_fit, gamma1_fit, r_fit = params_fit
gamma2_fit = 3.0 * gamma1_fit + r_fit**2

# fitted curve from params_fit_log (model expects log-space p)
y_fit = model_constrained(params_fit_log, t)

# --- 4. Stability loop (Monte Carlo initializations) ---
nTests = 100
results = np.zeros((nTests, 3))      # [k, gamma1, gamma2]
init_vals = np.zeros((nTests, 3))    # [k0, gamma1_0, r0] (physical)
RMSE_vals = np.zeros(nTests)

rng = np.random.default_rng()  # modern RNG

for i in range(nTests):
    k0_phys = k_init * (0.1 + (10.0 - 0.1) * rng.random())
    gamma1_0_phys = gamma1_init * (0.1 + (10.0 - 0.1) * rng.random())
    r0_phys = r_init * (0.5 + rng.random())

    params_init_log = np.log([k0_phys, gamma1_0_phys, r0_phys])
    init_vals[i, :] = [k0_phys, gamma1_0_phys, r0_phys]

    res_i = minimize(cost_fun, params_init_log, method='Nelder-Mead', options={'maxiter': 10000, 'disp': False})
    params_fit_i_log = res_i.x
    params_fit_i = np.exp(params_fit_i_log)

    k_i, gamma1_i, r_i = params_fit_i
    gamma2_i = 3.0 * gamma1_i + r_i**2
    results[i, :] = [k_i, gamma1_i, gamma2_i]

    y_fit_i = model_constrained(params_fit_i_log, t)
    RMSE_vals[i] = np.sqrt(np.mean((y_noisy - y_fit_i)**2))

# --- 5. Robust selection of "best" fit ---
RMSE_best_raw = RMSE_vals.min()
tol_factor = 1.5
valid_mask = RMSE_vals <= tol_factor * RMSE_best_raw

if not np.any(valid_mask):
    warnings.warn("No valid fits found within tolerance; using the single best RMSE.")
    idx_best = RMSE_vals.argmin()
else:
    median_params = np.median(results[valid_mask, :], axis=0)
    dist_to_median = np.sqrt(np.sum((results - median_params)**2, axis=1))
    valid_indices = np.where(valid_mask)[0]
    rel_positions = dist_to_median[valid_indices]
    idx_rel = rel_positions.argmin()
    idx_best = valid_indices[idx_rel]

best_params_fit = results[idx_best, :]   # physical params [k, gamma1, gamma2]
best_params_init = init_vals[idx_best, :]  # physical initial [k0, gamma1_0, r0]
RMSE_best = RMSE_vals[idx_best]

print("\n--- Robust best fit (RMSE-median filtered) ---")
print(f"RMSE = {RMSE_best:.6f}")
print(f"Initial parameters: k0={best_params_init[0]:.4f}, gamma1_0={best_params_init[1]:.4f}, r0={best_params_init[2]:.4f}")
print(f"Fitted parameters:  k={best_params_fit[0]:.4f}, gamma1={best_params_fit[1]:.4f}, gamma2={best_params_fit[2]:.4f}")

# reconstruct best log params for plotting
p_log_best = np.log(best_params_fit)  # [log(k), log(gamma1), log(r)] where third is log(r) because best_params_fit third stores gamma2 in MATLAB code?
# NOTE: In the original MATLAB you stored results(:,3) = gamma2 (not r). But best_params_fit is [k, gamma1, gamma2].
# To rebuild log-space p vector we need r, not gamma2. We must recover r from gamma2 = 3*gamma1 + r^2:
k_b = best_params_fit[0]
gamma1_b = best_params_fit[1]
gamma2_b = best_params_fit[2]
r_b = np.sqrt(max(gamma2_b - 3.0 * gamma1_b, 1e-12))
p_log_best = np.log([k_b, gamma1_b, r_b])

y_fit2 = model_constrained(p_log_best, t)

print("\n--- Javad fit (values from MAT file) ---")
print(f"k_Jav       = {k_var}")
print(f"gamma1_Jav  = {gamma1}")
print(f"gamma2_Jav  = {gamma2}")

# --- 6. Visualizations of stability: scatter plots linking initial -> estimated
gamma2_init_vals = 3.0 * init_vals[:, 1] + init_vals[:, 2]**2

fig1, axes = plt.subplots(1, 3, figsize=(18, 5))
axes[0].scatter(init_vals[:, 0], results[:, 0], marker='o')
axes[0].set_xlabel('k initial'); axes[0].set_ylabel('k estimated'); axes[0].set_title('k estimated vs k initial')
axes[0].plot([init_vals[:, 0].min(), init_vals[:, 0].max()], [best_params_fit[0], best_params_fit[0]], 'r--', linewidth=1.5)

axes[1].scatter(init_vals[:, 1], results[:, 1], marker='o')
axes[1].set_xlabel('gamma1 initial'); axes[1].set_ylabel('gamma1 estimated'); axes[1].set_title('gamma1 estimated vs gamma1 initial')
axes[1].plot([init_vals[:, 1].min(), init_vals[:, 1].max()], [best_params_fit[1], best_params_fit[1]], 'r--', linewidth=1.5)

axes[2].scatter(gamma2_init_vals, results[:, 2], marker='o')
axes[2].set_xlabel('gamma2 initial'); axes[2].set_ylabel('gamma2 estimated'); axes[2].set_title('gamma2 estimated vs gamma2 initial')
axes[2].plot([gamma2_init_vals.min(), gamma2_init_vals.max()], [best_params_fit[2], best_params_fit[2]], 'r--', linewidth=1.5)

fig1.suptitle('Stability of the nonlinear fit vs initial values', fontsize=14)
plt.tight_layout(rect=[0, 0.03, 1, 0.95])

# --- 7. Linear fit on second half of curve ---
N = t.size
start_idx = int(round(N/2.0))  # MATLAB used round(N/2):end
t_half = t[start_idx:]
y_half = y_noisy[start_idx:]
p_lin = np.polyfit(t_half, y_half, 1)
a_fit = p_lin[0]
b_fit = p_lin[1]

print("\n--- Linear fit results (second half) ---")
print(f"a (slope) = {a_fit:.6f}")
print(f"b (intercept) = {b_fit:.6f}")

# --- 8. Compare linear slope vs theoretical slopes ---
# The MATLAB code used:
# pente_theorique_V1 = 1/best_params_fit(3) + 1/best_params_fit(2)
# pente_theorique_V2 = 1/best_params_fit(3)
pente_theorique_V1 = 1.0 / best_params_fit[2] + 1.0 / best_params_fit[1]
pente_theorique_V2 = 1.0 / best_params_fit[2]

Ecart_V1 = 100.0 * abs(a_fit - pente_theorique_V1) / abs(pente_theorique_V1)
Ecart_V2 = 100.0 * abs(a_fit - pente_theorique_V2) / abs(pente_theorique_V2)
Best_ecart = min(Ecart_V1, Ecart_V2)

if Best_ecart == Ecart_V1:
    pente_theorique = pente_theorique_V1
else:
    pente_theorique = pente_theorique_V2

print("\n--- Slope comparison experimental vs model ---")
print(f"Linear fit slope (a)           = {a_fit:.6f}")
print(f"Theoretical slope (chosen)     = {pente_theorique:.6f}")
print(f"Relative difference (percent)  = {Best_ecart:.2f} %")

# --- 9. Global visualization with average fit ---
k_mean = best_params_fit[0]
gamma1_mean = best_params_fit[1]
gamma2_mean = best_params_fit[2]
y_fit_mean = (1.0 / k_mean) * (1.0 - np.exp(-k_mean * t / gamma1_mean)) + t / gamma2_mean

fig2, ax2 = plt.subplots(figsize=(8, 6))
ax2.plot(t, y_noisy, 'r.', label='Experimental data')
ax2.plot(t, y_fit_mean, 'g-', linewidth=2, label='Optimal Fit')
ax2.plot(t_half, np.polyval(p_lin, t_half), 'k--', linewidth=2, label='Linear fit (second half)')
ax2.set_xlabel('t')
ax2.set_ylabel('y')
ax2.legend(loc='best')
ax2.set_title('Best fit')
ax2.grid(True)

# save figure as JPEG
output_file = os.path.join(root_path, f"{filename}_fit.jpeg")
fig2.savefig(output_file, dpi=300)
print(f"\nSaved fit figure to: {output_file}")

# --- 10. Fit quality metrics (nonlinear)
residus = y_noisy - y_fit
RMSE = np.sqrt(np.mean(residus**2))
R2 = 1.0 - np.sum(residus**2) / np.sum((y_noisy - np.mean(y_noisy))**2)

print("\n--- Nonlinear fit quality ---")
print(f"RMSE = {RMSE:.6f}")
print(f"R^2  = {R2:.6f}")

# --- 11. Save results to Excel (append if file exists) ---
df = pd.DataFrame([{
    'filename': filename,
    'k': best_params_fit[0],
    'gamma1': best_params_fit[1],
    'gamma2': best_params_fit[2],
    'k_Jav': k_var,
    'gamma1_Jav': gamma1,
    'gamma2_Jav': gamma2
}])

excel_file = os.path.join(root_path, f"{filename}_FitParam.xlsx")
try:
    if os.path.isfile(excel_file):
        # append as a new row (openpyxl required)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # read existing sheet name (default: Sheet1) approach: append to same sheet
            # pandas' append via ExcelWriter is a bit cumbersome; simpler is write with startrow
            # We'll compute startrow by reading existing file
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            sheet_name = wb.sheetnames[0]
            # determine next row (including header)
            existing = pd.read_excel(excel_file, sheet_name=sheet_name)
            startrow = existing.shape[0] + 1
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
    else:
        df.to_excel(excel_file, index=False)
    print(f"Saved fit parameters to Excel: {excel_file}")
except Exception as e:
    warnings.warn(f"Could not append to Excel file due to: {e}. Writing new file instead.")
    df.to_excel(excel_file, index=False)
    print(f"Wrote fit parameters to Excel (new file): {excel_file}")

plt.show()
